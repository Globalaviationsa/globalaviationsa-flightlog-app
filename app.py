import os
import pandas as pd
from datetime import datetime, date, time
from openpyxl import load_workbook
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, send_file, redirect, url_for, flash

# ==== CONFIG ====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_XLSX = os.path.join(BASE_DIR, "FORMATTED TEMPLATE.xlsx")  # <-- UPDATED NAME
SHEET_NAME = None

START_ROW = 2  # first row of data (row 1 = header)

# Column positions
COL_DATE      = 1   # A
COL_DEP       = 2   # B
COL_DEP_TIME  = 3   # C   <-- SUM required
COL_ARR       = 4   # D
COL_ARR_TIME  = 5   # E
COL_AC_TYPE   = 7   # G
COL_REG       = 8   # H
COL_STUDENT   = 13  # M
COL_INSTR     = 14  # N
COL_FTYPE     = 15  # O
COL_PHASE     = 16  # P

# Time columns Y–AF (25–32)
COL_VFR       = 25  # Y
COL_IFR       = 26  # Z
COL_DAY       = 27  # AA
COL_NIGHT     = 28  # AB
COL_LOCAL     = 29  # AC
COL_XC        = 30  # AD
COL_PF        = 31  # AE
COL_PM        = 32  # AF

# SUM columns: C, I, J, K, L, Q–AF
SUM_COLS = [3, 9, 10, 11, 12] + list(range(17, 33))

TIME_COLS_MAPPING = {
    COL_VFR: "vfr_time",
    COL_IFR: "ifr_time",
    COL_DAY: "day_time",
    COL_NIGHT: "night_time",
    COL_LOCAL: "local_time",
    COL_XC: "cross_country_time",
    COL_PF: "pilot_flying_time",
    COL_PM: "pilot_monitoring_time",
}

REQUIRED_COLS = [
    "date", "departure", "arrival", "off_block", "on_block",
    "aircraft_type", "aircraft", "students", "instructors",
    "flight_type", "program_phase",
    "vfr_time", "ifr_time", "day_time", "night_time", "local_time",
    "cross_country_time", "pilot_flying_time", "pilot_monitoring_time",
]

app = Flask(__name__)
app.secret_key = "replace-with-any-secret"


# ==== HELPERS ====

def to_decimal_hours(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, time):
        return v.hour + v.minute/60 + v.second/3600
    if isinstance(v, datetime):
        t = v.time()
        return t.hour + t.minute/60 + t.second/3600

    s = str(v).strip()
    if s == "":
        return None

    if ":" in s:
        parts = s.split(":")
        try:
            h = int(parts[0]); m = int(parts[1])
            sec = int(parts[2]) if len(parts) > 2 else 0
            return h + m/60 + sec/3600
        except Exception:
            pass

    try:
        return float(s)
    except Exception:
        return None


def clean_instructor(name):
    if not isinstance(name, str):
        return name
    parts = name.strip().split()
    return parts[0] if parts else name


def parse_date_value(v):
    """Your CSV uses YYYY-MM-DD."""
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except:
        return None


def process_csv_to_excel(csv_path, output_path):
    raw = pd.read_csv(csv_path)

    missing = [c for c in REQUIRED_COLS if c not in raw.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    n_rows = len(raw)

    if not os.path.exists(TEMPLATE_XLSX):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_XLSX}")

    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active if SHEET_NAME is None else wb[SHEET_NAME]

    max_template_row = ws.max_row
    max_template_col = ws.max_column

    # Fill rows
    for i in range(max_template_row - START_ROW + 1):
        excel_row = START_ROW + i

        if i < n_rows:
            r = raw.iloc[i]

            # Date
            d = parse_date_value(r["date"])
            c = ws.cell(row=excel_row, column=COL_DATE)
            if d:
                c.value = d
                c.number_format = "dd/mm/yyyy"
            else:
                c.value = str(r["date"])

            ws.cell(row=excel_row, column=COL_DEP).value      = r["departure"]
            ws.cell(row=excel_row, column=COL_DEP_TIME).value = r["off_block"]
            ws.cell(row=excel_row, column=COL_ARR).value      = r["arrival"]
            ws.cell(row=excel_row, column=COL_ARR_TIME).value = r["on_block"]

            ws.cell(row=excel_row, column=COL_AC_TYPE).value = r["aircraft_type"]
            ws.cell(row=excel_row, column=COL_REG).value     = r["aircraft"]

            ws.cell(row=excel_row, column=COL_STUDENT).value = r["students"]
            ws.cell(row=excel_row, column=COL_INSTR).value   = clean_instructor(r["instructors"])

            ws.cell(row=excel_row, column=COL_FTYPE).value   = r["flight_type"]
            ws.cell(row=excel_row, column=COL_PHASE).value   = r["program_phase"]

            # Decimal times
            for col_idx, field in TIME_COLS_MAPPING.items():
                ws.cell(row=excel_row, column=col_idx).value = to_decimal_hours(r[field])

        else:
            # Clear unused rows
            for col_idx in (
                COL_DATE, COL_DEP, COL_DEP_TIME, COL_ARR, COL_ARR_TIME,
                COL_AC_TYPE, COL_REG, COL_STUDENT, COL_INSTR,
                COL_FTYPE, COL_PHASE,
                COL_VFR, COL_IFR, COL_DAY, COL_NIGHT,
                COL_LOCAL, COL_XC, COL_PF, COL_PM
            ):
                ws.cell(row=excel_row, column=col_idx).value = None

    # Find last data row
    last_data_row = 0
    for r in range(START_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_DATE).value not in (None, ""):
            last_data_row = r

    if last_data_row == 0:
        last_data_row = START_ROW

    total_row = last_data_row + 1

    # FILTER ONLY DATA ROWS (totals row excluded!)
    if ws.auto_filter:
        ws.auto_filter.ref = f"A1:AF{last_data_row}"

    # Totals row SUM formulas
    for col_idx in SUM_COLS:
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx).value = (
            f"=SUM({col_letter}{START_ROW}:{col_letter}{last_data_row})"
        )

    # Clear rows after totals
    for r in range(total_row + 1, ws.max_row + 1):
        for c in range(1, max_template_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.border = Border()

    # Print area
    ws.print_area = f"A1:AF{total_row}"

    wb.save(output_path)


# ==== ROUTES ====

@app.route("/", methods=["GET"])
def index():
    return render_template("upload.html")


@app.route("/convert", methods=["POST"])
def convert():
    if "file" not in request.files:
        flash("No file part in request.")
        return redirect(url_for("index"))

    file = request.files["file"]
    if file.filename == "":
        flash("No file selected.")
        return redirect(url_for("index"))

    upload_dir = os.path.join(BASE_DIR, "uploads")
    os.makedirs(upload_dir, exist_ok=True)

    csv_path = os.path.join(upload_dir, file.filename)
    file.save(csv_path)

    out_name = os.path.splitext(file.filename)[0] + "_formatted.xlsx"
    out_path = os.path.join(upload_dir, out_name)

    try:
        process_csv_to_excel(csv_path, out_path)
    except Exception as e:
        flash(str(e))
        return redirect(url_for("index"))

    return send_file(out_path, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)

